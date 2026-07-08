#!/usr/bin/env python3
"""Rigenera data.js dal file Excel della lega.

Uso:  python3 aggiorna_dati.py FantaNBA.xlsx
Requisiti:  pip install openpyxl
"""
import json
import re
import sys

from openpyxl import load_workbook

XLSX = sys.argv[1] if len(sys.argv) > 1 else "FantaNBA.xlsx"
SEASONS = ["2025/26", "2026/27", "2027/28", "2028/29", "2029/30"]
# salary cap per stagione (2025/26 ... 2029/30)
CAPS = [215_000_000, 230_000_000, 230_000_000, 230_000_000, 230_000_000]
CAP = CAPS[0]
NON_TEAM_SHEETS = {"Contratti", "Recap", "Bacheca", "Foglio10"}
ANNO_MIN_PICK = 2026   # pick mostrate dal draft di quest'anno in poi
RE_PICK = re.compile(r"scelta\s*(\d+)\s*°?\s*giro\s+(.+?)\s+((?:19|20)\d{2})", re.I)

wb = load_workbook(XLSX, read_only=True, data_only=True)


def estrai_pick(rows):
    """Scelte al draft future (dedup, ordinate) dalle righe di un foglio squadra."""
    trovate, viste = [], set()
    for row in rows:
        for v in row:
            if not isinstance(v, str) or "scelta" not in v.lower():
                continue
            m = RE_PICK.search(" ".join(v.replace("�", "°").split()))
            if not m:
                continue
            rd, origine, anno = int(m.group(1)), m.group(2).strip(), int(m.group(3))
            if anno < ANNO_MIN_PICK or (rd, origine.upper(), anno) in viste:
                continue
            viste.add((rd, origine.upper(), anno))
            trovate.append({"y": anno, "rd": rd, "from": origine})
    trovate.sort(key=lambda p: (p["y"], p["rd"], p["from"]))
    return trovate

# ---- Foglio Contratti ----
ws = wb["Contratti"]
rows = list(ws.iter_rows(values_only=True))
hdr = next(i for i, r in enumerate(rows) if r and r[0] == "Giocatore")

players = []
for r in rows[hdr + 1:]:
    if not r or not r[0]:
        continue
    players.append({
        "n": str(r[0]).strip(),
        "t": str(r[1]).strip() if r[1] else "",
        "r": str(r[2]).strip() if r[2] else "",
        "s": str(r[3]).strip() if r[3] else "",
        "sal": [float(r[4 + k] or 0) for k in range(5)],
        "opt": [str(r[9 + k]).strip() if len(r) > 9 + k and r[9 + k] else "" for k in range(5)],
    })

# ---- Fogli squadra (metadati) ----
teams = []
for sn in wb.sheetnames:
    if sn in NON_TEAM_SHEETS:
        continue
    ws = wb[sn]
    allrows = list(ws.iter_rows(values_only=True))
    vals = allrows[:9]

    def find(label):
        for row in vals:
            for c in row:
                if isinstance(c, str) and c.strip().upper().startswith(label):
                    return c.split(":", 1)[1].strip() if ":" in c else c
        return ""

    name = next((c.strip() for c in vals[1] if isinstance(c, str) and c.strip()), "")
    teams.append({
        "sheet": sn,
        "name": name,
        "gm": find("GM"),
        "city": find("CITY"),
        "franchise": find("UOMO FRANCHIGIA"),
        "nomina": find("ANNO NOMINA"),
        "scadenza": find("SCADENZA NOMINA"),
        "picks": estrai_pick(allrows),
    })

data = {
    "league": "FantaHezonja Champions",
    "cap": CAP,
    "caps": CAPS,
    "seasons": SEASONS,
    "teams": teams,
    "players": players,
}

# conserva i dati gestiti dal sito (rinnovi "pnd" dei giocatori e scelte al
# draft delle squadre): il sito è la fonte di queste informazioni, quindi
# vincono su quanto estratto ora dall'Excel.
try:
    prev = open("data.js", encoding="utf-8").read()
    prev = json.loads(re.search(r"window\.LEAGUE = (\{.*\});", prev, re.S).group(1))

    pnd_by_name = {p["n"]: p.get("pnd") for p in prev.get("players", []) if p.get("pnd")}
    kept = 0
    for p in players:
        if p["n"] in pnd_by_name:
            p["pnd"] = pnd_by_name[p["n"]]
            kept += 1
    if kept:
        print(f"Rinnovi da confermare conservati per {kept} giocatori.")

    # le pick sono gestite dall'Admin del sito: se in data.js ci sono, tienile
    picks_by_sheet = {t.get("sheet"): t.get("picks") for t in prev.get("teams", [])
                      if t.get("picks")}
    kept_p = 0
    for t in teams:
        if t.get("sheet") in picks_by_sheet:
            t["picks"] = picks_by_sheet[t["sheet"]]
            kept_p += 1
    if kept_p:
        print(f"Scelte al draft conservate dal sito per {kept_p} squadre.")
except (FileNotFoundError, AttributeError, json.JSONDecodeError):
    pass

payload = "window.LEAGUE = " + json.dumps(data, ensure_ascii=False) + ";"

# 1) file dati separato (comodo per modifiche a mano su GitHub)
with open("data.js", "w", encoding="utf-8") as f:
    f.write(payload)

# 2) aggiorna anche i dati incorporati in index.html, se presente
try:
    html = open("index.html", encoding="utf-8").read()
    html, n = re.subn(r"window\.LEAGUE = \{.*?\};", payload, html, count=1, flags=re.S)
    if n:
        open("index.html", "w", encoding="utf-8").write(html)
        print("index.html aggiornato con i nuovi dati.")
except FileNotFoundError:
    pass

print(f"OK: {len(players)} giocatori, {len(teams)} squadre.")
